# import 宣言
from collections import UserString
import openpyxl
import yaml
import logging
import sys
from rich.logging import RichHandler

# ストリームハンドラの設定とファイルハンドラの設定
rich_handler: RichHandler = RichHandler(rich_tracebacks=True)
rich_handler.setFormatter(logging.Formatter("%(asctime)s:%(lineno)d:%(levelname)s:%(message)s"))

# ファイルハンドラの設定(ログファイル出力先を設定)
logfile = logging.FileHandler('converte2y.log')
logfile.setFormatter(logging.Formatter("%(asctime)s:%(lineno)d:%(levelname)s:%(message)s"))

# ログレベルの設定とハンドラの関連づけ
logging.basicConfig(level=logging.INFO, handlers=[rich_handler,logfile])

# Logger名の指定
logger = logging.getLogger(__name__)

#変数設定
WorkDirPath = 'C:/Users/Administrator/Documents/ansible-python/2/lab/'
XcelFileName = 'AnsiblePythonLab.xlsx'

# main 
def main():

  # Excel ファイルを指定する
  wbook = openpyxl.load_workbook(WorkDirPath + XcelFileName)

  # Excel ファイルのシートリストを取得する
  sheets = wbook.sheetnames
  logger.debug("SheetList:"+str(sheets))

  # sheet の枚数分繰り返す
  for sheetname in sheets:
    logger.info("-- Processing:" + sheetname + "sheet --")
  
    # 処理するシートの取得
    sheet = wbook[sheetname]
    hostname = sheet["C6"]
    logging.debug("hostname=" + hostname.value)
    ip = sheet["C7"]
    logging.debug("ip=" + ip.value)
    cider = sheet["C8"]
    logging.debug("cider=" + str(cider.value))
    gateway = sheet["C9"]
    logging.debug("gateway=" + gateway.value)
    dns = sheet["C10"]
    logging.debug("dns=" + dns.value)

    packages = list()
    
    # インストールパッケージリストの取得
    packages = getlist(sheet,row_no=15,col_no=2,max_row_no=27)
    logger.debug(packages)

    # 起動/有効化サービスリストの取得
    services = getlist(sheet,row_no=15,col_no=5,max_row_no=27)
    logger.debug(services)

    # セキュリティファイアウォール(インバウンド許可リスト)の取得
    firewalls = getlist(sheet,row_no=30,col_no=2,max_row_no=35)
    logger.debug(firewalls)

    # 追加ユーザリストの取得
    users = getlist(sheet,row_no=40,col_no=2,max_row_no=47)
    logger.debug(users)

    # 追加ユーザリストのコメント取得
    users_comment = getlist(sheet,row_no=40,col_no=3,max_row_no=47)
    logger.debug(users_comment)

    # yaml 変換用複合ユーザリストの作成
    users_list = list()
    for (user,comment) in zip(users,users_comment):
      user_tmp = {'name':user,'comment':comment}
      users_list.append(user_tmp)
    logger.info(users_list)

    # yaml 出力データの整形
    out_data = {
      'hostanme': hostname.value,
      'ip': ip.value,
      'cider': cider.value,
      'gateway': gateway.value,
      'dns': dns.value,
      'packages': packages,
      'services': services,
      'firewalls': firewalls,
      'users': users_list,
      }
    logger.info(out_data)
  
    # 出力ファイルの指定
    yaml_output = open(WorkDirPath + "host_vars/" + sheetname + ".yaml",'w')

    # YAML へ変換
    yout = yaml.dump(out_data)
    
    # 各シートの情報を yaml に変換して出力
    yaml_output.write(yout)
    yaml_output.close()

    # yaml出力内容の出力
    logger.debug(yout)

  # man() 正常終了

# シートリストの取得関数
def getlist(sheet,row_no,col_no,max_row_no):
  convlist = list()

  # シートリストの処理
  while row_no < max_row_no:
    cell_value = sheet.cell(row=row_no,column=col_no).value
    if cell_value == None:
      row_no = row_no + 1
      continue
    convlist.append(cell_value)
    row_no = row_no + 1
  
  return convlist

# call main Function
if __name__ == '__main__':
  logger.info('START MAIN PROCESS')
  main()
